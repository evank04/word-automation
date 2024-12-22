import openpyxl
from docxtpl import DocxTemplate
import datetime
from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
import threading

file_name = ""
crane_name = ""


def strike(data: dict) -> dict:
    for cel, value in data.items():
        if value == 'Yes':
            data[cel] = 'Y / \u0336N\u0336'
        elif value == 'No':
            data[cel] = '\u0336Y\u0336 / N'
        elif value == 'Yes.':
            data[cel] = 'Yes'
        elif value == 'No.':
            data[cel] = 'No'

    return data


# wb = openpyxl.load_workbook(r'C:\Users\Evan\Desktop\proj\Excel Data\datass243.xlsx')
def process():
    global file_name
    wb = openpyxl.load_workbook(f'C:\\Users\\Evan\\Desktop\\proj\\Excel Data\\{file_name}')

    worksheet_info = {
        'FPC': ['fpc_data', 'A'],
        'OFPC': ['ofpc_data', 'B'],
        'SSC': ['ssc_data', 'C'],
        'CLPS': ['clps_data', 'D'],
        'CLODS': ['clods_data', 'E']
    }

    rows_to_process = {
        'FPC': ['B5:H6', 'B11:H12', 'B17:H18', 'B23:H24', 'B29:H30', 'B35:H36'],
        'OFPC': ['B5:H6', 'B11:H12', 'B17:H18', 'B23:H24', 'B29:H30', 'B35:H36', 'B41:H42'],
        'SSC': ['B5:F5', 'B11:H13'],
        'CLPS': ['B6:M9', 'B15:M18', 'B24:M37'],
        'CLODS': ['B6:M9', 'B15:M18', 'B24:M37']
    }

    datas = {key[0]: {} for key in worksheet_info.values()}

    for sheet_name, data_key in worksheet_info.items():
        ws = wb[sheet_name]
        for cell_range in rows_to_process[sheet_name]:
            for row in ws[cell_range]:
                for cell in row:
                    cell_value = cell.value
                    if cell_value is None:
                        cell_value = '\u200e'
                    datas[data_key[0]][data_key[1] + cell.coordinate] = cell_value

    doc = DocxTemplate(r'C:\Users\Evan\Desktop\proj\Base Template\sat_template_2.docx')

    final_data = {}

    for sheet_name in datas:
        processed_data = strike(datas[sheet_name])
        final_data.update(processed_data)

    final_data['crane'] = crane_name

    doc.render(final_data)
    time = datetime.datetime.now().strftime('%d_%m_%y')
    filepath = f'C:\\Users\\Evan\\Desktop\\proj\\Converted Report\\{final_data['crane']}_{time}.docx'
    doc.save(filepath)


def convert():
    global crane_name
    crane_name = crane.get().upper()
    bar = Progressbar(window, orient=HORIZONTAL, length=330, mode="indeterminate")
    bar.place(x=70, y=210)
    bar.start(10)

    def run_process():
        process()
        bar.stop()
        window.quit()

    threading.Thread(target=run_process).start()


def openfile():
    global file_name
    filename = filedialog.askopenfilename(initialdir="C:\\Users\\Evan\\Desktop\\proj")
    filename = filename.split("/")[-1]
    file_name = filename
    excel.delete(0, END)
    excel.insert(0, file_name)
    return filename


window = Tk()
window.geometry("420x240")
window.resizable(False, False)
window.title("SAT Data Entry")
window.iconphoto(True, PhotoImage(file='mvizn.png'))
window.config(background='#D9D9D9')

title = Label(window, text="Excel to Word Converter", font=('Arial', 16, 'bold'), background='#d9d9d9')
title.pack(pady=28)

crane_name = Label(window, text="Crane Name:", background='#d9d9d9')
crane_name.place(x=70, y=90)
crane = Entry(window)
crane.place(x=150, y=90)

excel_name = Label(window, text="Excel File:", background='#d9d9d9')
excel_name.place(x=70, y=135)
excel = Entry(window)
excel.place(x=150, y=135)

button = Button(text="Browse", command=openfile)
button.place(x=280, y=132)

convert_button = Button(window, text="Convert", command=convert)
convert_button.pack(side=BOTTOM, pady=40)

window.mainloop()
